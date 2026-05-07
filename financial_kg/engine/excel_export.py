"""全量 Excel 导出: 用 snapshot 值覆盖原始 Excel。"""
from __future__ import annotations

import os
import re
from datetime import datetime, date
from typing import Any

import openpyxl

# Match ISO datetime strings: "2023-12-31", "2023-12-31T00:00:00", "2023-12-31T00:00:00.000000"
_ISO_DATETIME_RE = re.compile(
    r"^(\d{4})-(\d{2})-(\d{2})(?:T(\d{2}):(\d{2}):(\d{2})(?:\.(\d+))?)?$"
)


def _sanitize_for_excel(value: Any) -> Any:
    """将 snapshot 中的值转换为 Excel 可识别的类型。

    - ISO 日期字符串 → datetime.date / datetime.datetime
    - 其他值保持原样
    """
    if isinstance(value, str):
        m = _ISO_DATETIME_RE.match(value)
        if m:
            year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if m.group(4):  # Has time component
                hour = int(m.group(4))
                minute = int(m.group(5))
                second = int(m.group(6))
                return datetime(year, month, day, hour, minute, second)
            return date(year, month, day)
    return value


def export_modified_excel(
    original_excel_path: str,
    snapshot_values: dict[str, Any],
    output_path: str,
    *,
    formula_cell_ids: set[str] | None = None,
) -> str:
    """加载原始 Excel, 用 snapshot 值覆盖对应单元格, 保留格式/合并/样式。

    Args:
        original_excel_path: 原始 .xlsx 文件路径
        snapshot_values: cell_id -> value 映射
        output_path: 输出文件路径
        formula_cell_ids: 公式单元格 ID 集合。传入时跳过这些单元格, 保留原公式。

    Returns:
        output_path
    """
    wb = openpyxl.load_workbook(original_excel_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                cell_id = f"{sheet_name}_{cell.row}_{col_letter}"
                if cell_id in snapshot_values:
                    if formula_cell_ids and cell_id in formula_cell_ids:
                        continue  # 保留原公式, 不覆盖
                    new_val = _sanitize_for_excel(snapshot_values[cell_id])
                    cell.value = new_val

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


def find_original_excel(task_id: str, output_dir: str) -> str | None:
    """查找原始 Excel 文件。

    优先查找 output/{task_id}_original.xlsx, 否则查找 output/ 下唯一的 .xlsx 文件。
    """
    candidate = os.path.join(output_dir, f"{task_id}_original.xlsx")
    if os.path.exists(candidate):
        return candidate

    xlsx_files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]
    if len(xlsx_files) == 1:
        return os.path.join(output_dir, xlsx_files[0])

    return None
